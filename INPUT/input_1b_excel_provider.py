"""  
INPUT-1B — Excel Worklist Provider

Purpose
-------
Read Excel workbooks and produce deterministic
worklists for runtime execution.

Provides workbook ingestion, sheet selection,
header resolution, ID normalization, and
optional manifest generation.

Public API
----------
extract_ids_from_excel(...)
load_worklist_ids(...)
iter_worklist_ids(...)
write_manifest_jsonl(...)
build_manifest_from_excel(...)

Dependencies
------------
openpyxl

Architecture Position
---------------------
RUN-1A
    ↓
PIPE-1E
    ↓
PIPE-1A
    ↓
INPUT-1B
    ↓
LOOP-1B

Status
------
Audited

Responsibilities
----------------
- Read Excel workbooks
- Resolve worksheet selection
- Resolve ID column headers
- Normalize ID values
- Remove duplicate IDs
- Produce runtime worklists
- Optionally generate manifest artifacts

Execution Flow
--------------
Workbook
    ↓
Sheet Resolution
    ↓
Header Resolution
    ↓
ID Extraction
    ↓
Normalization
    ↓
Worklist

Normalization Rules
-------------------
- Trim whitespace
- Convert numeric IDs to strings
- Remove blank values
- Preserve insertion order
- Optionally de-duplicate IDs

Security Notes
--------------
- Never log secrets
- Treat IDs as operational identifiers
- Performs no workflow execution
- Performs no audit logging
"""  
  
from __future__ import annotations  
  
import json  
from dataclasses import dataclass  
from pathlib import Path  
from typing import Iterable, Optional, Union  
  
__all__ = [  
    "ExcelIdSpec",  
    "extract_ids_from_excel",  
    "load_worklist_ids",  
    "iter_worklist_ids",  
    "write_manifest_jsonl",  
    "build_manifest_from_excel",  
    # smoke-test compatibility names expected by dev/dev_smoke_state_input.py  
    "read_ids_from_excel",  
    "excel_to_ids",  
    "dev_smoke",  
]  
  
_PathLike = Union[str, Path]  
  
  
def _coerce_path(p: _PathLike) -> Path:  
    return p if isinstance(p, Path) else Path(str(p))  
  
  
def _require_openpyxl_load_workbook():  
    # Lazy import so module import doesn't fail (prevents smoke test from "missing entry function")  
    try:  
        from openpyxl import load_workbook  # type: ignore  
    except Exception as e:  
        raise RuntimeError(  
            "openpyxl is required for INPUT-1B Excel reading.\n"  
            "Install it with: pip install openpyxl\n"  
            f"Import error: {e}"  
        )  
    return load_workbook  
  
  
@dataclass(frozen=True)  
class ExcelIdSpec:  
    """  
    Defines where IDs live in the Excel workbook.  
    - sheet: sheet name (e.g. "locations")  
    - header: the column header to pull IDs from (case-insensitive)  
    """  
  
    sheet: str = "locations"  
    header: str = "ACCOUNT_ID"  
  
  
def _resolve_sheet_name(sheetnames: list[str], requested: str) -> str:  
    """  
    Deterministic, safe fallback rules:  
    1) exact match requested  
    2) case-insensitive match requested  
    3) sheet named "Worklist" (case-insensitive)  
    4) if workbook has exactly 1 sheet, use it  
    else: raise KeyError  
    """  
    if requested in sheetnames:  
        return requested  
  
    req_l = requested.strip().lower()  
    for s in sheetnames:  
        if s.strip().lower() == req_l:  
            return s  
  
    for s in sheetnames:  
        if s.strip().lower() == "worklist":  
            return s  
  
    if len(sheetnames) == 1:  
        return sheetnames[0]  
  
    raise KeyError(f"Sheet not found: {requested!r}. Available: {sheetnames}")  
  
  
def extract_ids_from_excel(  
    xlsx_path: _PathLike,  
    *,  
    spec: ExcelIdSpec = ExcelIdSpec(),  
    allow_numeric: bool = True,  
    drop_blanks: bool = True,  
    dedupe: bool = True,  
) -> list[str]:  
    """  
    Read an Excel workbook and extract ID values from a named sheet + header column.  
  
    Rules:  
    - Finds header row by scanning row 1 for the requested header (case-insensitive).  
    - Reads down until the end of the used range.  
    - Converts numeric IDs to int->str when allow_numeric=True (so 1001.0 becomes "1001").  
    - Trims whitespace, optionally drops blanks, optionally de-dupes preserving order.  
    """  
    load_workbook = _require_openpyxl_load_workbook()  
  
    path = _coerce_path(xlsx_path).expanduser().resolve()  
    if not path.exists():  
        raise FileNotFoundError(f"Excel file not found: {path}")  
  
    wb = load_workbook(filename=path, read_only=True, data_only=True)  
    try:  
        sheet_name = _resolve_sheet_name(list(wb.sheetnames), spec.sheet)  
        ws = wb[sheet_name]  
  
        target_header = spec.header.strip().lower()  
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)  
        if not header_row:  
            raise ValueError(f"Sheet {sheet_name!r} has no header row.")  
  
        col_idx: Optional[int] = None  
        for i, v in enumerate(header_row):  
            if v is None:  
                continue  
            if str(v).strip().lower() == target_header:  
                col_idx = i  # 0-based  
                break  
  
        if col_idx is None:  
            raise KeyError(  
                f"Header {spec.header!r} not found in sheet {sheet_name!r}. "  
                f"Headers seen: {[str(x).strip() if x is not None else '' for x in header_row]}"  
            )  
  
        out: list[str] = []  
        seen: set[str] = set()  
  
        for row in ws.iter_rows(min_row=2, values_only=True):  
            raw = row[col_idx] if col_idx < len(row) else None  
  
            if raw is None:  
                s = ""  
            elif allow_numeric and isinstance(raw, (int, float)):  
                if isinstance(raw, float) and raw.is_integer():  
                    s = str(int(raw))  
                else:  
                    s = str(raw).strip()  
            else:  
                s = str(raw).strip()  
  
            if drop_blanks and not s:  
                continue  
  
            if dedupe:  
                if s in seen:  
                    continue  
                seen.add(s)  
  
            out.append(s)  
  
        return out  
    finally:  
        try:  
            wb.close()  
        except Exception:  
            pass  
  
  
# ---- INPUT-1B config style wrappers ----  
def load_worklist_ids(cfg: dict) -> list[str]:  
    """  
    INPUT-1B style: loads IDs for the run.  
  
    Expected cfg keys:  
      - WORKLIST_XLSX (or INPUT_XLSX): path to workbook  
      - WORKLIST_SHEET: optional (default "locations")  
      - WORKLIST_HEADER: optional (default "ACCOUNT_ID")  
    """  
    xlsx = cfg.get("WORKLIST_XLSX") or cfg.get("INPUT_XLSX")  
    if not xlsx:  
        raise ValueError("Missing cfg['WORKLIST_XLSX'] (or INPUT_XLSX)")  
  
    sheet = str(cfg.get("WORKLIST_SHEET", "locations"))  
    header = str(cfg.get("WORKLIST_HEADER", "ACCOUNT_ID"))  
    return extract_ids_from_excel(xlsx, spec=ExcelIdSpec(sheet=sheet, header=header))  
  
  
def iter_worklist_ids(cfg: dict) -> Iterable[str]:  
    """Optional generator form if your PIPE prefers iterators."""  
    yield from load_worklist_ids(cfg)  
  
  
def write_manifest_jsonl(  
    *,  
    ids: Iterable[str],  
    manifest_path: _PathLike,  
    manifest_key_field: str,  
) -> int:  
    """  
    Write a minimal manifest JSONL file with one object per ID:  
      { <manifest_key_field>: "<id>" }  
  
    Returns number of records written.  
    """  
    mp = _coerce_path(manifest_path)  
    mp.parent.mkdir(parents=True, exist_ok=True)  
  
    count = 0  
    with mp.open("w", encoding="utf-8", newline="\n") as fh:  
        for wid in ids:  
            wid_norm = str(wid).strip()  
            if not wid_norm:  
                continue  
            fh.write(json.dumps({manifest_key_field: wid_norm}, ensure_ascii=False) + "\n")  
            count += 1  
    return count  
  
  
def build_manifest_from_excel(  
    *,  
    excel_path: _PathLike,  
    sheet_name: str,  
    key_column: str,  
    manifest_key_field: str,  
    manifest_path: _PathLike,  
) -> int:  
    """  
    Convenience wrapper: extract IDs from Excel and write a minimal manifest JSONL.  
  
    Returns number of records written.  
    """  
    ids = extract_ids_from_excel(  
        excel_path,  
        spec=ExcelIdSpec(sheet=sheet_name, header=key_column),  
    )  
    return write_manifest_jsonl(  
        ids=ids,  
        manifest_path=manifest_path,  
        manifest_key_field=manifest_key_field,  
    )  
  
  
# ---- Smoke-test compatibility wrappers (names dev/dev_smoke_state_input.py tries) ----  
def read_ids_from_excel(excel_path: _PathLike, sheet_name: str, key_column: str) -> list[str]:  
    return extract_ids_from_excel(  
        excel_path,  
        spec=ExcelIdSpec(sheet=sheet_name, header=key_column),  
    )  
  
  
def excel_to_ids(  
    *,  
    excel_path: _PathLike,  
    sheet_name: str,  
    key_column: str,  
    manifest_path: Optional[_PathLike] = None,  
    manifest_key_field: Optional[str] = None,  
) -> list[str]:  
    ids = read_ids_from_excel(excel_path, sheet_name, key_column)  
    if manifest_path is not None:  
        mk = str(manifest_key_field or key_column)  
        write_manifest_jsonl(ids=ids, manifest_path=manifest_path, manifest_key_field=mk)  
    return ids  
  
  
def dev_smoke() -> None:  
    from tempfile import TemporaryDirectory  
  
    try:  
        from openpyxl import Workbook  # type: ignore  
    except Exception as e:  
        raise RuntimeError(  
            "openpyxl is required to run INPUT-1B dev_smoke().\n"  
            "Install it with: pip install openpyxl\n"  
            f"Import error: {e}"  
        )  
  
    with TemporaryDirectory(prefix="input_1b_excel_provider_smoke_") as td:  
        tmp = Path(td)  
        xlsx = tmp / "input_1b_excel_provider_smoke.xlsx"  
  
        wb = Workbook()  
        ws = wb.active  
        ws.title = "Worklist"  
        ws["A1"] = "ACCOUNT_ID"  
        ws["A2"] = "  1001  "  
        ws["A3"] = 1002  # numeric -> "1002"  
        ws["A4"] = ""  
        wb.save(xlsx)  
  
        # Request a missing sheet; should fall back to "Worklist"  
        ids = extract_ids_from_excel(xlsx, spec=ExcelIdSpec(sheet="locations", header="ACCOUNT_ID"))  
        assert ids == ["1001", "1002"]  