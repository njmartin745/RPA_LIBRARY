"""  
Smoke test for top-level INPUT-1B shim: input_1b_excel_provider.py  
  
How to run:  
  python dev/dev_smoke_input_1b_excel_provider.py  
"""  
  
from __future__ import annotations  
  
import json  
import tempfile  
import traceback  
from pathlib import Path  
import sys  
  
ROOT = Path(__file__).resolve().parents[1]  
if str(ROOT) not in sys.path:  
    sys.path.insert(0, str(ROOT))  
  
from INPUT.input_1b_excel_provider import (
    excel_to_ids,  
    read_ids_from_excel,  
)  
  
  
def _read_jsonl(path: Path) -> list[dict]:  
    out: list[dict] = []  
    with path.open("r", encoding="utf-8") as fh:  
        for ln in fh:  
            ln = ln.strip()  
            if not ln:  
                continue  
            out.append(json.loads(ln))  
    return out  
  
  
def _make_excel(path: Path, sheet_name: str, key_column: str) -> None:  
    try:  
        from openpyxl import Workbook  # type: ignore  
    except Exception as e:  
        raise RuntimeError(  
            "openpyxl is required for this smoke test.\n"  
            "Install it with: pip install openpyxl\n"  
            f"Import error: {e}"  
        )  
  
    path.parent.mkdir(parents=True, exist_ok=True)  
    wb = Workbook()  
    ws = wb.active  
    ws.title = sheet_name  
  
    ws.cell(row=1, column=1, value=key_column)  
    values = ["100", "200", None, "   ", "300"]  
    for i, v in enumerate(values, start=2):  
        ws.cell(row=i, column=1, value=v)  
  
    wb.save(path)  
  
  
def main() -> int:  
    try:  
        with tempfile.TemporaryDirectory(prefix="dev_smoke_input_1b_") as td:  
            d = Path(td)  
            excel_path = d / "locations.xlsx"  
            manifest_path = d / "manifest.jsonl"  
  
            _make_excel(excel_path, sheet_name="locations", key_column="key_ID")  
  
            ids = read_ids_from_excel(excel_path, "locations", "key_ID")  
            assert ids == ["100", "200", "300"]  
  
            ids2 = excel_to_ids(  
                excel_path=excel_path,  
                sheet_name="locations",  
                key_column="key_ID",  
                manifest_path=manifest_path,  
                manifest_key_field="key_ID",  
            )  
            assert ids2 == ["100", "200", "300"]  
            assert manifest_path.exists()  
  
            recs = _read_jsonl(manifest_path)  
            assert recs == [{"key_ID": "100"}, {"key_ID": "200"}, {"key_ID": "300"}]  
  
            print("PASS: dev_smoke_input_1b_excel_provider")  
            return 0  
  
    except Exception as e:  
        print("FAIL: dev_smoke_input_1b_excel_provider")  
        print(f"Error: {type(e).__name__}: {e}")  
        traceback.print_exc()  
        return 1  
  
  
if __name__ == "__main__":  
    raise SystemExit(main())  