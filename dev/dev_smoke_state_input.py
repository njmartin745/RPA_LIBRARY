# dev_smoke_state_input.py
"""
Smoke test for:
- INPUT/input_1b_excel_provider.py
- STATE/state_1b_manifest_jsonl.py
 
What it does:
1) Creates a tiny Excel file (if missing) with sheet+column you can control.
2) Uses INPUT-1B to extract IDs and (optionally) write a baseline manifest.jsonl.
3) Uses STATE-1B to choose active manifest, load IDs, and append an audit record.
4) Creates a retry manifest with a subset of IDs and verifies STATE chooses it.
 
How to run:
  python dev_smoke_state_input.py
 
Notes:
- Adjust imports at the top if your function names differ.
- This script is intentionally "loud" with prints and assertions.
"""
 
from __future__ import annotations
 
import json
from pathlib import Path
import sys
 
ROOT = Path(__file__).resolve().parents[1]

# Ensure repo root is importable when running: python dev/dev_smoke_state_input.py  
_REPO_ROOT = Path(__file__).resolve().parents[1]  
if str(_REPO_ROOT) not in sys.path:  
    sys.path.insert(0, str(_REPO_ROOT))  

if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
    
from typing import List, Optional
 
# ---- Adjust these imports to match your module public APIs ----
# INPUT  
excel_to_ids = None  # type: ignore  
read_ids_from_excel = None  # type: ignore  
build_manifest_from_excel = None  # type: ignore  
  
_input_import_errors: list[str] = []  
  
for modname in ("input_1b_excel_provider", "INPUT.input_1b_excel_provider"):  
    try:  
        m = __import__(modname, fromlist=["*"])  
        excel_to_ids = getattr(m, "excel_to_ids", None) or excel_to_ids  
        read_ids_from_excel = getattr(m, "read_ids_from_excel", None) or read_ids_from_excel  
        build_manifest_from_excel = getattr(m, "build_manifest_from_excel", None) or build_manifest_from_excel  
  
        if any([excel_to_ids, read_ids_from_excel, build_manifest_from_excel]):  
            print(f"OK: INPUT-1B imported from: {getattr(m, '__file__', modname)}")  
            break  
        else:  
            _input_import_errors.append(  
                f"{modname}: imported but missing expected functions "  
                f"(has excel_to_ids={hasattr(m,'excel_to_ids')}, "  
                f"read_ids_from_excel={hasattr(m,'read_ids_from_excel')}, "  
                f"build_manifest_from_excel={hasattr(m,'build_manifest_from_excel')})"  
            )  
    except Exception as e:  
        _input_import_errors.append(f"{modname}: {type(e).__name__}: {e}")  
 
# STATE  
try:  
    from STATE.state_1b_manifest_jsonl import (  # type: ignore  
        append_jsonl_line,  
        choose_active_manifest,  
        load_ids_from_manifest,  
        utc_ts,  
        write_audit,  
    )  
except Exception:  
    # Legacy fallback (if someone keeps STATE modules at repo root)  
    from state_1b_manifest_jsonl import (  # type: ignore  
        append_jsonl_line,  
        choose_active_manifest,  
        load_ids_from_manifest,  
        utc_ts,  
        write_audit,  
    )  
 
# ----------------------------------------------------------------
 
 
def _require_one(*fns):
    """Return the first non-None function; fail with helpful message otherwise."""
    for fn in fns:
        if fn is not None:
            return fn
    raise RuntimeError(
        "Could not find a usable INPUT-1B entry function. "
        "Update the imports at the top of dev_smoke_state_input.py to match your INPUT module.\n"
        "Tried: excel_to_ids, read_ids_from_excel, build_manifest_from_excel"
    )
 
 
def _ensure_test_excel(
    excel_path: Path, sheet_name: str, key_column: str
) -> None:
    """
    Create a tiny Excel file if it doesn't exist.
    Uses openpyxl directly to avoid pandas dependency in the smoke test.
    """
    if excel_path.exists():
        return
 
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception as e:
        raise RuntimeError(
            "openpyxl is required to auto-create the test Excel file.\n"
            "Install it with: pip install openpyxl\n"
            f"Import error: {e}"
        )
 
    excel_path.parent.mkdir(parents=True, exist_ok=True)
 
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
 
    # header
    ws.cell(row=1, column=1, value=key_column)
 
    # rows (include blanks / whitespace)
    values = ["100", "200", None, "   ", "300"]
    for i, v in enumerate(values, start=2):
        ws.cell(row=i, column=1, value=v)
 
    wb.save(excel_path)
 
 
def _read_jsonl_lines(path: Path) -> List[dict]:
    out: List[dict] = []
    with path.open("r", encoding="utf-8") as fh:
        for ln in fh:
            ln = ln.strip()
            if not ln:
                continue
            out.append(json.loads(ln))
    return out
 
 
def main() -> int:
    # --- Config for the smoke test ---
    excel_path = Path("input/locations_smoke.xlsx")
    sheet_name = "locations"
    key_column = "key_ID"
 
    data_dir = Path("data_smoke")
    baseline_manifest = data_dir / "manifest.jsonl"
    retry_manifest = data_dir / "manifest_retry.jsonl"
    audit_log = data_dir / "rpa_audit.jsonl"
 
    manifest_key_field = "key_ID"
 
    print("== Smoke Test: INPUT-1B + STATE-1B ==")
    print(f"Excel:   {excel_path}")
    print(f"Sheet:   {sheet_name}")
    print(f"Column:  {key_column}")
    print(f"Baseline manifest: {baseline_manifest}")
    print(f"Retry manifest:    {retry_manifest}")
    print(f"Audit log:         {audit_log}")
    print()
 
    # Clean prior smoke outputs
    if data_dir.exists():
        for p in [baseline_manifest, retry_manifest, audit_log]:
            if p.exists():
                p.unlink()
    data_dir.mkdir(parents=True, exist_ok=True)
 
    # 1) Ensure Excel exists
    _ensure_test_excel(excel_path, sheet_name, key_column)
    assert excel_path.exists(), "Excel file should exist after ensure step."
    print("OK: test Excel present")
 
    if excel_to_ids is None and read_ids_from_excel is None and build_manifest_from_excel is None:  
        raise RuntimeError("INPUT-1B import failed:\n- " + "\n- ".join(_input_import_errors))  

    # 2) Run INPUT to extract IDs (and optionally write baseline manifest)
    input_fn = _require_one(excel_to_ids, read_ids_from_excel, build_manifest_from_excel)
 
    ids: Optional[List[str]] = None
 
    # Try a couple call signatures based on common patterns.
    # If your function is different, update this section.
    try:
        # Pattern: excel_to_ids(excel_path, sheet_name, key_column, ...)
        ids = input_fn(
            excel_path=excel_path,
            sheet_name=sheet_name,
            key_column=key_column,
            manifest_path=baseline_manifest,
            manifest_key_field=manifest_key_field,
        )
    except TypeError:
        try:
            # Pattern: read_ids_from_excel(excel_path, sheet_name, key_column)
            ids = input_fn(excel_path, sheet_name, key_column)
        except TypeError:
            # Pattern: build_manifest_from_excel returns count; then we load IDs via STATE
            count = input_fn(
                excel_path=excel_path,
                manifest_path=baseline_manifest,
                sheet_name=sheet_name,
                key_column=key_column,
                manifest_key_field=manifest_key_field,
            )
            print(f"INPUT wrote baseline manifest lines: {count}")
            ids = None
 
    if isinstance(ids, int):
        print(f"INPUT wrote baseline manifest lines: {ids}")
        ids = None
    elif isinstance(ids, list):
        print(f"INPUT returned IDs: {ids}")
    else:
        print(f"INPUT returned unexpected type: {type(ids)}")
 
    # Ensure baseline manifest exists (either written by INPUT or you can write it here)
    assert baseline_manifest.exists(), (
        "Baseline manifest.jsonl was not created. "
        "Either adjust INPUT-1B to write it, or update this smoke test to write it."
    )
 
    # Sanity-check baseline manifest contents
    baseline_records = _read_jsonl_lines(baseline_manifest)
    assert len(baseline_records) == 3, "Baseline manifest should have 3 records"
    assert all(manifest_key_field in r for r in baseline_records), "Each record should have manifest_key_field"
    print("OK: baseline manifest looks valid")
 
    # 3) STATE chooses active manifest (no retry yet -> baseline)
    active = choose_active_manifest(
        baseline_manifest=baseline_manifest,
        retry_manifest=retry_manifest,
    )
    assert active == baseline_manifest, "Expected baseline as active when retry doesn't exist/empty"
    print("OK: choose_active_manifest -> baseline")
 
    # 4) STATE loads IDs from manifest
    loaded_ids = load_ids_from_manifest(active, manifest_key_field=manifest_key_field)
    print(f"STATE loaded IDs: {loaded_ids}")
    assert loaded_ids == ["100", "200", "300"], "STATE load_ids_from_manifest mismatch"
    print("OK: load_ids_from_manifest -> expected IDs")
 
    # 5) Append an audit record via STATE
    write_audit(
        audit_log_path=audit_log,
        record={"timestamp": utc_ts(), "event": "SMOKE_START", "count": len(loaded_ids)},
        swallow_errors=False,
    )
    assert audit_log.exists(), "Audit log should exist after write_audit"
    audit_records = _read_jsonl_lines(audit_log)
    assert len(audit_records) == 1, "Audit log should have 1 record"
    print("OK: write_audit appended 1 record")
 
    # 6) Create a retry manifest with a subset and verify active switches
    append_jsonl_line(retry_manifest, {manifest_key_field: "200"})
    active2 = choose_active_manifest(
        baseline_manifest=baseline_manifest,
        retry_manifest=retry_manifest,
    )
    assert active2 == retry_manifest, "Expected retry as active when retry has remaining IDs"
    print("OK: choose_active_manifest -> retry")
 
    loaded_retry_ids = load_ids_from_manifest(active2, manifest_key_field=manifest_key_field)
    assert loaded_retry_ids == ["200"], "Expected retry IDs to be ['200']"
    print("OK: load_ids_from_manifest(retry) -> ['200']")
 
    print()
    print("✅ SMOKE TEST PASSED")
    print(f"- Baseline manifest: {baseline_manifest}")
    print(f"- Retry manifest:    {retry_manifest}")
    print(f"- Audit log:         {audit_log}")
 
    return 0
 
 
if __name__ == "__main__":
    raise SystemExit(main())